from __future__ import annotations

import io
import json
import shutil
import subprocess
import tempfile
from datetime import date as date_type
from datetime import timedelta
from pathlib import Path

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import render
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

from orgs.decorators import require_department_roles
from orgs.models import DepartmentMember

from handover.models import HandoverItem, HandoverSession


def _parse_date_yyyy_mm_dd(s: str) -> date_type:
    try:
        y, m, d = s.split("-")
        return date_type(int(y), int(m), int(d))
    except Exception as e:
        raise Http404("日期格式错误") from e


def _tri_status_label(v: str) -> str:
    mapping = {
        "yes": "√",
        "no": "×",
        "other": "其他",
    }
    return mapping.get((v or "").strip(), (v or "").strip() or "-")


def _resolve_soffice_path() -> str | None:
    configured = getattr(settings, "ORS_SOFFICE_PATH", None)
    if configured:
        p = Path(configured)
        if p.is_file():
            return str(p)
    for name in ("soffice", "soffice.exe"):
        found = shutil.which(name)
        if found:
            return found
    candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    return None


def _pdf_dependency_error(message: str) -> HttpResponse:
    """配置/依赖问题不应使用 404，避免与「链接不存在」混淆。"""
    return HttpResponse(message, status=503, content_type="text/plain; charset=utf-8")


def _apply_daily_handover_sheet(ws, department, target_date: date_type, session: HandoverSession):
    headers = [
        "科室",
        "姓名",
        "年龄",
        "手术名称",
        "特殊病情交接",
        "输血前九项",
        "压疮评估单",
        "皮肤情况",
        "术前访视",
        "特殊器械准备",
    ]

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    col_count = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(row=1, column=1, value=f"{department.name}交班报告").font = title_font
    ws.cell(row=1, column=1).alignment = header_alignment
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    ws.cell(
        row=2,
        column=1,
        value=(
            f"日期 {target_date}；"
            f"择期 {session.elective_count or 0} 台；急诊 {session.emergency_count or 0} 台；"
            f"抢救/特殊 {session.rescue_count or 0} 台；"
            f"备注：{(session.notes or '').strip() or '无'}；"
            f"交班人 _________；接班人 _________"
        ),
    ).alignment = left_alignment
    ws.row_dimensions[2].height = 24

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_count)
    ws.cell(
        row=3,
        column=1,
        value=(
            f"标本交接：{_tri_status_label(session.specimen_handover_status)} {session.specimen_handover_note or ''}；"
            f"层流运行：{_tri_status_label(session.laminar_flow_running_status)} {session.laminar_flow_running_note or ''}；"
            f"生物监测：{_tri_status_label(session.bio_monitoring_status)} {session.bio_monitoring_note or ''}；"
            f"急救车：{_tri_status_label(session.crash_cart_status)} {session.crash_cart_note or ''}"
        ),
    ).alignment = left_alignment
    ws.row_dimensions[3].height = 24

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=col_count)
    ws.cell(
        row=4,
        column=1,
        value=(
            f"消防安全：{_tri_status_label(session.fire_safety_status)} {session.fire_safety_note or ''}；"
            f"钥匙管理：{_tri_status_label(session.key_management_status)} {session.key_management_note or ''}；"
            f"其他合格证收纳：{_tri_status_label(session.certs_in_place_status)} {session.certs_in_place_note or ''}；"
            f"其它事件：{(session.other_incidents or '').strip() or '无'}"
        ),
    ).alignment = left_alignment
    ws.row_dimensions[4].height = 24

    ws.append(headers)
    header_row = 5
    ws.row_dimensions[5].height = 24
    for col in range(1, col_count + 1):
        c = ws.cell(row=header_row, column=col)
        c.font = header_font
        c.alignment = header_alignment
        c.border = border

    first_data_row = header_row + 1
    for item in session.items.order_by("id"):
        ws.append(
            [
                item.department_text,
                item.patient_name,
                item.age,
                item.surgery_name,
                item.special_handover,
                item.blood_transfusion_checks,
                item.pressure_ulcer_assessment,
                item.skin_condition,
                item.preop_visit,
                item.special_instruments,
            ]
        )

    if ws.max_row >= first_data_row:
        for row in ws.iter_rows(min_row=first_data_row, max_row=ws.max_row, min_col=1, max_col=col_count):
            for cell in row:
                if cell.column in (4, 5, 10):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=col_count):
        for cell in row:
            if cell.border != border:
                cell.border = border

    ws.freeze_panes = f"A{first_data_row}"

    widths = [12, 12, 6, 24, 24, 12, 12, 12, 10, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = None
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.25, bottom=0.25, header=0.1, footer=0.1)
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_area = f"A1:J{ws.max_row}"


@login_required
@require_department_roles(DepartmentMember.Role.ADMIN)
def export_handover_excel(request, dept_code: str, date: str):
    department = request.department  # type: ignore[attr-defined]
    target_date = _parse_date_yyyy_mm_dd(date)

    session = HandoverSession.objects.prefetch_related("items").filter(department=department, handover_date=target_date).first()
    if not session:
        raise Http404("当日交班不存在")

    wb = Workbook()
    ws = wb.active
    ws.title = f"{target_date}"
    _apply_daily_handover_sheet(ws=ws, department=department, target_date=target_date, session=session)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"handover_{department.code or department.id}_{target_date}.xlsx"
    return FileResponse(buf, as_attachment=True, filename=filename)


@login_required
@require_department_roles(DepartmentMember.Role.ADMIN)
def export_handover_pdf(request, dept_code: str, date: str):
    department = request.department  # type: ignore[attr-defined]
    target_date = _parse_date_yyyy_mm_dd(date)

    session = HandoverSession.objects.prefetch_related("items").filter(department=department, handover_date=target_date).first()
    if not session:
        raise Http404("当日交班不存在")

    soffice = _resolve_soffice_path()
    if not soffice:
        return _pdf_dependency_error(
            "未检测到 LibreOffice，无法执行 Excel 转 PDF。"
            "请安装 LibreOffice（https://www.libreoffice.org/），"
            "或将 soffice 加入系统 PATH，或在 .env 中设置 ORS_SOFFICE_PATH 为 soffice.exe 的完整路径。"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = f"{target_date}"
    _apply_daily_handover_sheet(ws=ws, department=department, target_date=target_date, session=session)

    with tempfile.TemporaryDirectory(prefix="ors_handover_") as temp_dir:
        temp_path = Path(temp_dir)
        src_xlsx = temp_path / f"handover_{target_date}.xlsx"
        out_pdf = temp_path / f"handover_{target_date}.pdf"
        wb.save(src_xlsx)
        try:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(temp_path), str(src_xlsx)],
                check=True,
                capture_output=True,
                timeout=120,
            )
        except Exception:
            return _pdf_dependency_error("Excel 转 PDF 失败，请检查 LibreOffice 安装与 ORS_SOFFICE_PATH 配置。")
        if not out_pdf.exists():
            return _pdf_dependency_error("PDF 文件生成失败（LibreOffice 未产出目标文件）。")
        pdf_bytes = out_pdf.read_bytes()

    filename = f"handover_{department.code or department.id}_{target_date}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_department_roles(DepartmentMember.Role.ADMIN)
def report_center(request, dept_code: str):
    department = request.department  # type: ignore[attr-defined]
    today = timezone.localdate()
    start_date = request.GET.get("start_date") or str(today - timedelta(days=6))
    end_date = request.GET.get("end_date") or str(today)
    start = _parse_date_yyyy_mm_dd(start_date)
    end = _parse_date_yyyy_mm_dd(end_date)
    if start > end:
        start, end = end, start

    sessions = (
        HandoverSession.objects.filter(department=department, handover_date__gte=start, handover_date__lte=end)
        .order_by("handover_date")
        .prefetch_related("items")
    )
    items = HandoverItem.objects.filter(session__in=sessions)
    total_sessions = sessions.count()
    total_items = items.count()

    date_labels = []
    date_counts = []
    for s in sessions:
        count = int(s.items.count())
        if count > 0:
            date_labels.append(str(s.handover_date))
            date_counts.append(count)
    top_surgeries = {}
    for it in items:
        name = (it.surgery_name or "").strip() or "未填写手术名称"
        top_surgeries[name] = top_surgeries.get(name, 0) + 1
    top_surgeries = sorted(((k, int(v)) for k, v in top_surgeries.items() if v > 0), key=lambda x: x[1], reverse=True)[:8]

    return render(
        request,
        "m/report_center.html",
        {
            "department": department,
            "dept_code": dept_code,
            "start_date": start,
            "end_date": end,
            "total_sessions": total_sessions,
            "total_items": total_items,
            "date_labels_json": json.dumps(date_labels, ensure_ascii=False),
            "date_counts_json": json.dumps(date_counts, ensure_ascii=False),
            "top_surgeries_json": json.dumps([{"name": x[0], "value": x[1]} for x in top_surgeries], ensure_ascii=False),
        },
    )


@login_required
@require_department_roles(DepartmentMember.Role.ADMIN)
def export_handover_range_excel(request, dept_code: str):
    department = request.department  # type: ignore[attr-defined]
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    if not start_date or not end_date:
        raise Http404("缺少日期范围")
    start = _parse_date_yyyy_mm_dd(start_date)
    end = _parse_date_yyyy_mm_dd(end_date)
    if start > end:
        start, end = end, start

    sessions = list(
        HandoverSession.objects.filter(department=department, handover_date__gte=start, handover_date__lte=end)
        .order_by("handover_date")
        .prefetch_related("items")
    )
    if not sessions:
        raise Http404("该时间段内没有交班记录")

    wb = Workbook()
    wb.remove(wb.active)
    for session in sessions:
        ws = wb.create_sheet(title=str(session.handover_date)[:31])
        _apply_daily_handover_sheet(
            ws=ws,
            department=department,
            target_date=session.handover_date,
            session=session,
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"handover_range_{department.code or department.id}_{start}_{end}.xlsx"
    return FileResponse(buf, as_attachment=True, filename=filename)


@login_required
@require_department_roles(DepartmentMember.Role.ADMIN)
def export_handover_range_pdf(request, dept_code: str):
    department = request.department  # type: ignore[attr-defined]
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    if not start_date or not end_date:
        raise Http404("缺少日期范围")
    start = _parse_date_yyyy_mm_dd(start_date)
    end = _parse_date_yyyy_mm_dd(end_date)
    if start > end:
        start, end = end, start

    sessions = list(
        HandoverSession.objects.filter(department=department, handover_date__gte=start, handover_date__lte=end)
        .order_by("handover_date")
        .prefetch_related("items")
    )
    if not sessions:
        raise Http404("该时间段内没有交班记录")

    soffice = _resolve_soffice_path()
    if not soffice:
        return _pdf_dependency_error(
            "未检测到 LibreOffice，无法执行 Excel 转 PDF。"
            "请安装 LibreOffice（https://www.libreoffice.org/），"
            "或将 soffice 加入系统 PATH，或在 .env 中设置 ORS_SOFFICE_PATH 为 soffice.exe 的完整路径。"
        )

    wb = Workbook()
    wb.remove(wb.active)
    for session in sessions:
        ws = wb.create_sheet(title=str(session.handover_date)[:31])
        _apply_daily_handover_sheet(
            ws=ws,
            department=department,
            target_date=session.handover_date,
            session=session,
        )

    with tempfile.TemporaryDirectory(prefix="ors_handover_range_") as temp_dir:
        temp_path = Path(temp_dir)
        src_xlsx = temp_path / f"handover_range_{start}_{end}.xlsx"
        out_pdf = temp_path / f"handover_range_{start}_{end}.pdf"
        wb.save(src_xlsx)
        try:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(temp_path), str(src_xlsx)],
                check=True,
                capture_output=True,
                timeout=180,
            )
        except Exception:
            return _pdf_dependency_error("区间 Excel 转 PDF 失败，请检查 LibreOffice 安装与 ORS_SOFFICE_PATH 配置。")
        if not out_pdf.exists():
            return _pdf_dependency_error("区间 PDF 文件生成失败（LibreOffice 未产出目标文件）。")
        pdf_bytes = out_pdf.read_bytes()

    filename = f"handover_range_{department.code or department.id}_{start}_{end}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


